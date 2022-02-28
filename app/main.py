import platform
import os
import sys
import math
import openpyxl
import datetime  # one of the functions doesnt work unless I have this line, idk
from datetime import date
import pandas as pd
from typing import List, Tuple, Set, Dict
from typing import Optional
from fastapi import FastAPI
from pydantic import BaseModel
app = FastAPI()

PATH: str
# This is a directory in which material files is stored.
directory: str = r'/code/app/MD04'
saftey_stock: int
stock: int
avg_stock_change: float


# This function is to find stock
def find_stock(date: datetime, data: pd.DataFrame(), formatted_date: str) -> int:
    data_stock = data[data[6] == "Stock"]
    for index, row in data_stock.iterrows():
        if date == row[2]:
            return row[8]

    # If else no concrete "Stock" data present just take stock for first entry of that day
    data_date = data[data[2] == date]
    for index, row in data_date.iterrows():
        # Assuming data sorted, returning first total_quantity entry for that data
        return row[8]

    # Else essentially
    print("No data found for", formatted_date)
    return None


def find_saftey_stock(date: datetime, data: pd.DataFrame(), saftey_stock: int) -> int:
    for index, row in data.iterrows():
        return row[7]  # Change due to removal of saftey stock, assumed const

    # Else
    return saftey_stock  # Return last known value of saftey stock


def format_date(date: datetime) -> str:
    plt = platform.system()

    # Not sure if any operating systems are missing or also need consideration
    if plt == "Windows":
        return datetime.date.strftime(date, "%#m/%#d/%Y")
    else:  # Unix
        return datetime.date.strftime(date, "%-m/%-d/%Y")


def calc_avg_stock_change(data: pd.DataFrame()) -> float:
    data_filtered = data[data[6] == "Stock"]  # Stock values
    # Assume data is sorted for now
    prev_row: int
    curr_row: int
    counter: int = 0
    total_dif: int = 0
    first_val_in = False
    for index, row in data_filtered.iterrows():
        if row[8] == "total_quantity":
            continue
        elif first_val_in == False:
            prev_row = row[8]
            first_val_in = True
        else:
            curr_row = row[8]
            diff = curr_row - prev_row
            if diff < 0:
                total_dif += diff
                counter += 1
                prev_row = curr_row
            else:
                prev_row = curr_row

    return total_dif/counter


def get_health_score(stock: int, saftey_stock: int, k_val: float) -> float:
    # Change value of k to affect attitude of curve
    if stock != None and saftey_stock != 0:
        return sigmoid(SS=(stock/saftey_stock), k=k_val)
    else:
        return None


def sigmoid(SS: int, k: float) -> float:
    return (((1/(1+math.exp(-k*SS)))-(1/2))*2)*100


def print_values(health: float, stock: int, avg_stock_change: float, material: str, formatted_date: str):
    if stock != None:
        DoS = abs(stock/avg_stock_change)
        print("Health score for", material, "on",
              formatted_date, "is", health, "DoS is", DoS)


# This API would list All materials
# This function will return a list of all material in MD04 directory
@app.get("/materials")
def material():
    materials_list: List = []
    mat_id: str
    desc: str
    desc_eng: str
    saftey_stock: int
    descriptionDirectory: str = r'/code/app/MD04/Desc'

    for file in os.listdir(descriptionDirectory):
        if file.endswith('.xlsx'):
            mat_id = file[5:len(file)-5]

            # Convert to .xlsx so it's easier to work with
            PATH = descriptionDirectory+"/"+"desc_"+mat_id+".xlsx"
            if not os.path.exists(PATH):
                read_file = pd.read_csv(
                    descriptionDirectory+"/"+file_csv)
                read_file.to_excel(PATH, index=None, header=True)

            wb = openpyxl.load_workbook(PATH)
            sheet = wb.active
            data = pd.DataFrame(sheet.values)
            for index, row in data.iterrows():
                if index == 0:
                    continue
                else:
                    desc = row[6]
                    desc_eng = row[7]
                    saftey_stock = row[19]
                    break

            materials = {
                'Material': mat_id,
                'Description': desc,
                'Description Eng': desc_eng,
                'Saftey Stock': saftey_stock,
            }

            materials_list.append(materials)

    return materials_list

    # This function will return a list of all material in MD04 directory
    # materials = {
    #     'Material':  "Material ID",
    #     'Description': "Material description",
    #     'Description Eng': "Material description eng",
    #     'Safety Stock': "Safety Stock",
    # }

    # return materials

# This API will return basic information of a material ID


@app.get("/material/{id}")
def materialInfo(id: str):
    desc: str
    desc_eng: str
    saftey_stock: int
    descriptionDirectory: str = r'/code/app/MD04/Desc'
    #PATH = descriptionDirectory+"/"+"desc_"+mat_id+".xlsx"

    PATH = descriptionDirectory+"/"+"desc_"+id+".xlsx"

    if not os.path.exists(PATH):
        for file_csv in os.listdir(directory):
            if file_csv.endswith('.csv') and mat_id in file_csv:
                read_file = pd.read_csv(directory+"/"+file_csv)
                read_file.to_excel(PATH, index=None, header=True)

    wb = openpyxl.load_workbook(PATH)
    sheet = wb.active
    data = pd.DataFrame(sheet.values)

    for index, row in data.iterrows():
        if index == 0:
            continue
        else:
            desc = row[6]
            desc_eng = row[7]
            saftey_stock = row[19]
            break

    materialInfo = {
        'Material': id,
        'Description': desc,
        'Description Eng': desc_eng,
        'Saftey Stock': saftey_stock,
    }

    return materialInfo

    # materialInfo = {
    #     'Material':  id,
    #     'Description': "Material description",
    #     'Description Eng': "Material description eng",
    #     'Safety Stock': "Safety Stock",
    # }

    # return materialInfo


class MaterialException(BaseModel):
    materialID: str
    date: str
    numDays: int
    limit: int

# This API will return exception information (System View)


@app.post("/exceptions")
def exceptionInfo(exceptions: MaterialException):

    # exceptionList will be a JSON list,
    # which contains sorted list (highest to lowest) of materials
    exceptionList = {
        'Material':  exceptions.materialID,
        'Date': exceptions.date,
        'TotalException': "It will be sum of all exceptions",
        'Percentage': "percentage with respect to the total exception",
        'NumDays': exceptions.numDays
    }

    return exceptionList


class Health(BaseModel):
    materialID: str
    date: str


@app.post("/healthscore")
def healthScore(healthscore: Health):

    material = healthscore.materialID
    date = healthscore.date
    num_days = 10

    # Get desired material file
    for file_name in os.listdir(directory):
        if material in file_name:
            PATH = directory+"/"+file_name
            break

    wb = openpyxl.load_workbook(PATH)
    sheet = wb.active
    data = pd.DataFrame(sheet.values)

    avg_stock_change: float = calc_avg_stock_change(data)

    # Start at given date find stock
    # If no stock present indicate that
    # Find saftey stock, if no number present use previous number or 0
    # Get health status
    # Repeat for next 10 days
    mm, dd, yyyy = map(int, date.split('/'))
    date_obj = datetime.datetime(yyyy, mm, dd)
    saftey_stock = 0  # Default
    avg: List = []  # List to keep track of health scores

    for i in range(int(num_days)):
        td = datetime.timedelta(days=i)
        new_date = date_obj + td
        formatted_date = format_date(date=new_date)

        # ASSUME DATA SORTED ALREADY
        stock = find_stock(new_date, data, formatted_date)
        saftey_stock = abs(find_saftey_stock(
            new_date, data[data[6] == "SafeSt"], saftey_stock))

        health = get_health_score(stock, saftey_stock, k_val=0.8)
        if health != None:
            avg.append(health)

        print_values(health, stock, avg_stock_change, material, formatted_date)

    result = sum(avg)/len(avg)
    result = round(result, 2)
    percentage_result = str(result).__add__(' %')

    health_score = {
        "Material": material,
        "Date": date,
        "Health-score": percentage_result,
        "Safety Stock": saftey_stock,
        "Status Code": 200,
        "Status Description": "Success"
    }

    return health_score
